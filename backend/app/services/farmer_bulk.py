"""Farmer Registration - Bulk Upload: template generation and sheet validation.

One COLUMNS spec drives both the template writer and the parser, so the file a District
Admin fills in can never drift from what the importer accepts.

Two things here are deliberately stricter than the rest of the app:

  * The sheet demands 24 fields where `FarmerIn` requires 7. The extra ones are enforced
    here rather than on the schema, because the Register Farmer dialog must keep working
    with its own (much smaller) required set.
  * The uploaded workbook is parsed in memory and never written to disk. Every other
    upload in this app becomes a FileRecord on the uploads volume; this one must not,
    because it carries raw Aadhaar numbers.
"""
import io
import re
from datetime import date, datetime
from typing import Any, Callable, Optional

from sqlalchemy.orm import Session

from app.core.aadhaar import aadhaar_hash, normalize_aadhaar
from app.models import (
    Activity, AssetType, Caste, District, EducationLevel, Farmer, Religion,
    Lac, SericultureCircle, SilkType, SilkTypeActivityProduct, User,
)
from app.routers.lands import VALID_LAND_TYPES

MAX_ROWS = 500
GENDERS = ["Male", "Female", "Other"]
FARMER_TYPES = ["Small", "Medium", "Large"]
YES_NO = ["Yes", "No"]

# Read-only, derived from the Sericulture Circle — never parsed, never stored.
LAC_KEY = "lac_display"

SHEET_FARMERS = "Farmers"
SHEET_REFERENCE = "Reference"
SHEET_INSTRUCTIONS = "Instructions"

# Row layout of the Farmers sheet. Row 3 carries the activity ids (hidden) so resolution
# never depends on header text — a renamed activity stays importable.
ROW_HEADER = 1
ROW_KEYS = 2
ROW_EXAMPLE = 3
ROW_FIRST_DATA = 4

AADHAAR_WARNING = (
    "This file will contain real Aadhaar numbers. The application itself never stores them "
    "in readable form, but this spreadsheet does. Keep it off shared drives and email, and "
    "delete it from your computer as soon as the upload has succeeded."
)


class Column:
    """One spreadsheet column.

    `source` names a Reference-sheet vocabulary when the cell is a dropdown; `None` means
    the admin types the value. `required` is the bulk-upload rule, which is stricter than
    the API's own.
    """

    def __init__(self, header: str, key: str, required: bool = False,
                 source: Optional[str] = None, example: str = "", note: str = ""):
        self.header = header
        self.key = key
        self.required = required
        self.source = source
        self.example = example
        self.note = note


# Column order is set by the District Admin's own reordered template and must not be
# changed casually — the sheet they fill in offline is the spec. The activity grid sits in
# the middle of the sheet, not at the end, so the writer expands it after this key.
ACTIVITY_BLOCK_AFTER = "family_member_female"

# `source` names a Reference-sheet vocabulary when the cell is a dropdown; None means typed.
COLUMNS: list[Column] = [
    Column("District", "district_name", True, "Districts", "", "Pre-filled — do not change"),
    Column("First Name", "first_name", True, None, "Rekha"),
    Column("Middle Name", "middle_name", False, None, "Kumari"),
    Column("Last Name", "last_name", True, None, "Das"),
    Column("Gender", "gender", True, "Genders", "Female"),
    Column("Date of Birth", "date_of_birth", True, None, "14-05-1988", "dd-mm-yyyy"),
    Column("Mobile", "mobile_no", True, None, "9101010101", "10 digits"),
    Column("Aadhaar", "aadhaar_no", True, None, "123456789012", "12 digits"),
    Column("PAN Number", "pan_no", False, None, "ABCDE1234F", "Optional — many farmers have none"),
    Column("Sericulture Circle", "seri_circle_id", True, "Circles", ""),
    # Not a farmer field. The app derives it from the circle (see frontend/src/lib/lac.ts),
    # so this column fills itself by formula and the parser ignores it entirely.
    Column("LAC", LAC_KEY, False, None, "",
           "Fills itself from the Sericulture Circle"),
    Column("Village", "village_name", True, None, "Barbari"),
    Column("Panchayat", "gaon_panchayat", True, None, "Barbari GP"),
    Column("Development Block", "development_block", True, None, "Chandrapur"),
    Column("Post Office", "post_office", True, None, "Chandrapur"),
    Column("PIN Code", "pin_code", True, None, "781021"),
    Column("Farmer Type", "farmer_type", True, "FarmerTypes", "Small"),
    Column("Education Level", "education_level_id", True, "EducationLevels", ""),
    Column("Caste", "caste_id", True, "Castes", "OBC"),
    Column("Religion", "religion_id", True, "Religions", "Hindu"),
    Column("Family Members (Male)", "family_member_male", True, None, "2"),
    Column("Family Members (Female)", "family_member_female", True, None, "3"),
    # ---- the 15 silk type / activity Yes/No columns are expanded here ----
    Column("Experience (years)", "experience_years", True, None, "12"),
    Column("Experience in Activities", "experience_activity_ids", True, None,
           "Eri > Eri Rearing; Muga > Muga Reeling", "Separate several with ;"),
    Column("Dag No", "dag_no", False, None, "114", "Land — optional"),
    Column("Patta No", "patta_no", False, None, "27", "Land — optional"),
    Column("Land Type", "land_type", False, "LandTypes", "Owned", "Defaults to Owned"),
    Column("Asset Type", "asset_type_id", False, "AssetTypes", "", "Asset — optional"),
    Column("Quantity", "asset_quantity", False, None, "1", "Defaults to 1"),
    Column("Year Acquired", "asset_year", False, None, "2021", "e.g. 2021"),
    Column("Account Number", "account_number", True, None, "30123456789"),
    Column("Bank Name", "bank_name", True, None, "State Bank of India"),
    Column("Branch Name", "branch_name", True, None, "Chandrapur"),
    Column("IFSC Code", "ifsc_code", True, None, "SBIN0007654"),
]

MANDATORY_HEADERS = [c.header for c in COLUMNS if c.required and c.key != "district_name"]
OPTIONAL_HEADERS = [c.header for c in COLUMNS if not c.required]


# --------------------------------------------------------------------------- vocabularies

def _activity_label(silk_name: str, activity_name: str) -> str:
    """`Eri > Eri Rearing`. Activity names are NOT unique on their own — 'Food Plant
    Plantation' exists once per silk type locally — so every reference is qualified."""
    return f"{silk_name} > {activity_name}"


def load_vocabularies(db: Session, district_id: str) -> dict:
    """Everything the dropdowns and the resolver need, read once."""
    silk = {s.id: s.silk_type_name for s in db.query(SilkType).all()}
    activities = db.query(Activity).filter(Activity.is_active).all()
    activities.sort(key=lambda a: (silk.get(a.silk_type_id, ""), a.step_no or 0))

    district = db.query(District).filter(District.id == district_id).first()
    circles = db.query(SericultureCircle).filter(
        SericultureCircle.district_id == district_id, SericultureCircle.is_active).all()
    # A FIG-level shared asset can never be owned by an individual farmer, so it is never
    # offered here — same rule AssetRowsEditor applies in the dialog.
    asset_types = [a for a in db.query(AssetType).filter(AssetType.is_active).all()
                   if a.ownership_level != "FIG"]
    # The LAC is derived from the circle, never stored on the farmer — this map backs both
    # the Reference lookup block and the worked example.
    lacs = {l.id: l.lac_name for l in db.query(Lac).all()}
    lac_by_circle = {c.circle_name: (lacs.get(c.lac_id) or "")
                     for c in sorted(circles, key=lambda c: c.circle_name)}

    return {
        "district": district,
        "district_name": district.district_name if district else "",
        "activities": [(a, _activity_label(silk.get(a.silk_type_id, "?"), a.activity_name))
                       for a in activities],
        "circles": {c.circle_name: c.id for c in sorted(circles, key=lambda c: c.circle_name)},
        "castes": {c.caste_name: c.id for c in db.query(Caste).filter(Caste.is_active).all()},
        "religions": {r.religion_name: r.id for r in db.query(Religion).filter(Religion.is_active).all()},
        "education": {e.education_level_name: e.id
                      for e in db.query(EducationLevel).filter(EducationLevel.is_active).all()},
        "asset_types": {a.name: a.id for a in sorted(asset_types, key=lambda a: a.name)},
        "lac_by_circle": lac_by_circle,
    }


def _examples(vocab: dict) -> dict[str, str]:
    """Worked-example values for the columns whose valid entries come from master data, so
    the example row is always something the dropdown will actually accept. The circle is
    picked as the first one mapped to an LAC, so the derived LAC cell demonstrates the
    lookup instead of sitting empty."""
    circles = list(vocab["circles"])
    linked = [c for c in circles if vocab["lac_by_circle"].get(c)]
    circle = (linked or circles or [""])[0]
    education = sorted(vocab["education"])
    assets = sorted(vocab["asset_types"])
    return {
        "district_name": vocab["district_name"],
        "seri_circle_id": circle,
        LAC_KEY: vocab["lac_by_circle"].get(circle) or "(circle not yet mapped to an LAC)",
        "education_level_id": education[0] if education else "",
        "asset_type_id": assets[0] if assets else "",
    }


def _reference_lists(vocab: dict) -> dict[str, list[str]]:
    return {
        "Districts": [vocab["district_name"]],
        "Genders": GENDERS,
        "FarmerTypes": FARMER_TYPES,
        "YesNo": YES_NO,
        "Circles": sorted(vocab["circles"]),
        "Castes": sorted(vocab["castes"]),
        "Religions": sorted(vocab["religions"]),
        "EducationLevels": sorted(vocab["education"]),
        "LandTypes": sorted(VALID_LAND_TYPES),
        "AssetTypes": sorted(vocab["asset_types"]),
        "Activities": [label for _, label in vocab["activities"]],
    }


# --------------------------------------------------------------------------- template

def build_template_xlsx(db: Session, district_id: str) -> bytes:
    """The workbook is generated from live master data, so it cannot offer a value the
    importer would then reject."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.worksheet.datavalidation import DataValidation

    vocab = load_vocabularies(db, district_id)
    refs = _reference_lists(vocab)
    activities = vocab["activities"]

    examples = _examples(vocab)
    wb = Workbook()

    # ---- Reference sheet first: the dropdowns point at named ranges defined on it ----
    ws_ref = wb.active
    ws_ref.title = SHEET_REFERENCE
    ws_ref["A1"] = "Accepted values — do not edit. The Farmers sheet reads from these."
    ws_ref["A1"].font = Font(bold=True)
    for idx, (name, values) in enumerate(refs.items()):
        col = get_column_letter(idx + 1)
        ws_ref[f"{col}3"] = name
        ws_ref[f"{col}3"].font = Font(bold=True)
        for offset, value in enumerate(values):
            ws_ref[f"{col}{4 + offset}"] = value
        if values:
            ref = f"'{SHEET_REFERENCE}'!${col}$4:${col}${3 + len(values)}"
            wb.defined_names.add(DefinedName(f"list_{name}", attr_text=ref))
        ws_ref.column_dimensions[col].width = max(len(name), *(len(v) for v in values or [""])) + 3

    # Two-column circle -> office block backing the SDO/CDC formula. Placed after the
    # single-column vocabularies; its position is computed, never hardcoded.
    # Only circles that actually have an office are listed: a matched-but-empty cell makes
    # Excel's VLOOKUP return 0, whereas no match at all falls through IFERROR to "".
    pairs = [(c, o) for c, o in vocab["lac_by_circle"].items() if o]
    off_first = get_column_letter(len(refs) + 1)
    off_last = get_column_letter(len(refs) + 2)
    ws_ref[f"{off_first}3"] = "Circle"
    ws_ref[f"{off_last}3"] = "LAC"
    for cell in (ws_ref[f"{off_first}3"], ws_ref[f"{off_last}3"]):
        cell.font = Font(bold=True)
    for offset, (circle_name, office) in enumerate(pairs):
        ws_ref[f"{off_first}{4 + offset}"] = circle_name
        ws_ref[f"{off_last}{4 + offset}"] = office
    ws_ref.column_dimensions[off_first].width = 24
    ws_ref.column_dimensions[off_last].width = 30
    office_range = (f"'{SHEET_REFERENCE}'!${off_first}$4:${off_last}"
                    f"${3 + max(len(pairs), 1)}")

    # ---- Farmers sheet ----
    ws = wb.create_sheet(SHEET_FARMERS, 0)
    mandatory_fill = PatternFill("solid", fgColor="D6E4D9")

    # The activity grid sits mid-sheet, so build the final ordered column list first
    # rather than appending activities at the end.
    ordered: list[tuple[Column, Optional[Activity]]] = []
    for col in COLUMNS:
        ordered.append((col, None))
        if col.key == ACTIVITY_BLOCK_AFTER:
            for activity, label in activities:
                ordered.append((Column(label, f"activity:{activity.id}", True, "YesNo"), activity))

    headers = [c.header for c, _ in ordered]
    circle_letter = None
    sdo_idx = None
    for idx, (col, activity) in enumerate(ordered, start=1):
        cell = ws.cell(row=ROW_HEADER, column=idx,
                       value=f"{col.header} *" if col.required else col.header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        if col.required:
            cell.fill = mandatory_fill
        ws.cell(row=ROW_KEYS, column=idx, value=col.key)
        if activity is not None:
            # A worked example must satisfy "at least one Yes", so the first one is ticked.
            ws.cell(row=ROW_EXAMPLE, column=idx,
                    value="Yes" if activity is activities[0][0] else "No")
            ws.column_dimensions[get_column_letter(idx)].width = 22
        else:
            ws.cell(row=ROW_EXAMPLE, column=idx,
                    value=examples.get(col.key, col.example))
            ws.column_dimensions[get_column_letter(idx)].width = min(max(len(col.header) + 4, 14), 32)
        if col.key == "seri_circle_id":
            circle_letter = get_column_letter(idx)
        if col.key == LAC_KEY:
            sdo_idx = idx

    ws.row_dimensions[ROW_KEYS].hidden = True
    ws.row_dimensions[ROW_HEADER].height = 34
    for cell in ws[ROW_EXAMPLE]:
        cell.font = Font(italic=True, color="8A8F87")
    ws.freeze_panes = ws.cell(row=ROW_FIRST_DATA, column=1)

    last_row = ROW_FIRST_DATA + MAX_ROWS
    for idx, (col, _) in enumerate(ordered, start=1):
        letter = get_column_letter(idx)
        # SDO/CDC is computed, so it gets a formula instead of a dropdown.
        if col.key == LAC_KEY:
            continue
        if col.source and (col.source == "YesNo" or refs.get(col.source)):
            dv = DataValidation(type="list", formula1=f"=list_{col.source}", allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{letter}{ROW_FIRST_DATA}:{letter}{last_row}")

    # Fills itself the moment a Circle is chosen — the same resolution the app does on
    # screen via sdoCdcName(). Row 3 keeps a literal value so the worked example is
    # visible even in a viewer that does not recalculate on open.
    if sdo_idx and circle_letter:
        for row in range(ROW_FIRST_DATA, last_row + 1):
            ws.cell(row=row, column=sdo_idx,
                    value=f'=IFERROR(VLOOKUP(${circle_letter}{row},{office_range},2,FALSE),"")')

    _write_instructions(wb, vocab, headers)
    wb.move_sheet(SHEET_REFERENCE, offset=1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_instructions(wb, vocab: dict, headers: list[str]) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    ws = wb.create_sheet(SHEET_INSTRUCTIONS)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 96

    def line(label: str, text: str = "", bold: bool = False):
        r = ws.max_row + 1
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(bold=True)
        ws[f"B{r}"] = text
        ws[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
        if bold:
            ws[f"B{r}"].font = Font(bold=True)
        return r

    ws["A1"] = "Farmer Registration - Bulk Upload"
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    line("District", f'{vocab["district_name"]} — pick it from the dropdown on every row.')
    line("How to use",
         "Fill one row per farmer on the Farmers sheet, starting at row 4. Row 3 is a greyed-out "
         "example — leave it, it is skipped on import. Upload the file from Farmer Management "
         "and review the preview before confirming; nothing is saved until you confirm.")
    line("Row limit", f"{MAX_ROWS} farmers per file.")
    ws.append([])

    r = line("AADHAAR — PLEASE READ", AADHAAR_WARNING, bold=True)
    ws[f"B{r}"].fill = PatternFill("solid", fgColor="F5DDDB")
    ws.row_dimensions[r].height = 46
    ws.append([])

    line("Mandatory", ", ".join(MANDATORY_HEADERS))
    line("Mandatory (at least one)",
         "One of the silk type / activity columns must be Yes — a farmer must produce something.")
    line("Optional", ", ".join(OPTIONAL_HEADERS))
    line("Not collected here",
         "Photo and Bank passbook are uploaded per farmer after import, from Edit Farmer. "
         "The Farmer Code is generated automatically. Person-with-disability is not recorded "
         "by this sheet.")
    line("Editing later",
         "Every farmer created by this upload can be edited afterwards from Farmer Management, "
         "exactly like one registered through the form — including adding more land parcels "
         "or assets than the single set this sheet carries.")
    ws.append([])
    line("Dropdown columns",
         "Gender, Sericulture Circle, Farmer Type, Education Level, Caste, Religion, Land Type, "
         "Asset Type and every activity column pick from the Reference sheet. Everything else "
         "is typed.")
    line("LAC",
         "Fills itself as soon as you choose the Sericulture Circle — leave it alone. It is "
         "shown for your reference only and is ignored on upload, because the application "
         "works it out from the circle. It stays blank for a circle that has not been mapped "
         "to a constituency yet in Master Data.")
    line("Date of Birth", "dd-mm-yyyy (a real Excel date cell also works).")
    line("Experience in Activities",
         "Copy the exact wording from the Reference sheet's Activities list, separating several "
         "with a semicolon — e.g. Eri > Eri Rearing; Muga > Muga Reeling")
    line("Land and Asset",
         "Both optional. Leave every land cell blank for no parcel; Land Type defaults to Owned. "
         "Leave Asset Type blank for no asset; Quantity defaults to 1. A farmer with more than "
         "one parcel or asset gets the rest added later from Edit Farmer.")
    line("Logins", "Every imported farmer gets a login on their mobile number with the standard "
                   "default password, which they should change on first sign-in.")


# --------------------------------------------------------------------------- parsing

_DATE_FORMATS = ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y")


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_date(value: Any) -> Optional[date]:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _clean(value)
    if not text:
        return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_int(value: Any) -> Optional[int]:
    text = _clean(value)
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


class RowErrors:
    """Collects every problem on a row. Nearly the whole sheet is mandatory, so stopping at
    the first would turn fixing a file into a dozen upload cycles."""

    def __init__(self, row_no: int):
        self.row_no = row_no
        self.messages: list[str] = []

    def add(self, message: str) -> None:
        self.messages.append(message)

    def __bool__(self) -> bool:
        return bool(self.messages)


def parse_and_validate(db: Session, district_id: str, file_bytes: bytes) -> tuple[list[dict], list[dict], list[str]]:
    """Returns (ready_rows, row_errors, sheet_errors). Reads only — writes nothing."""
    from openpyxl import load_workbook

    sheet_errors: list[str] = []
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception:
        return [], [], ["That file could not be read as an Excel workbook (.xlsx)."]
    if SHEET_FARMERS not in wb.sheetnames:
        return [], [], [f"The workbook has no '{SHEET_FARMERS}' sheet — please use the "
                        f"downloaded template."]

    ws = wb[SHEET_FARMERS]
    keys = [_clean(c.value) for c in ws[ROW_KEYS]]
    if not keys or keys[0] != "district_name":
        return [], [], ["This does not look like the Farmer Registration template — please "
                        "download a fresh copy and fill that in."]

    vocab = load_vocabularies(db, district_id)
    by_label = {label: activity for activity, label in vocab["activities"]}
    by_id = {activity.id: activity for activity, _ in vocab["activities"]}
    stap_by_activity = _output_staps_by_activity(db, list(by_id))

    known_mobiles = {m for (m,) in db.query(Farmer.mobile_no).all()}
    known_mobiles |= {m for (m,) in db.query(User.mobile_no).all()}
    known_aadhaar = {h for (h,) in db.query(Farmer.aadhaar_hash).all() if h}

    seen_mobiles: dict[str, int] = {}
    seen_aadhaar: dict[str, int] = {}
    ready: list[dict] = []
    errors: list[dict] = []

    # District is pre-filled all the way down the template, so a row is only "used" if
    # something OTHER than District was typed into it. Count those, not the row span,
    # or an untouched template reads as 500 rows.
    used: list[tuple[int, tuple]] = []
    for offset, values in enumerate(ws.iter_rows(min_row=ROW_FIRST_DATA, values_only=True)):
        cells = {keys[i]: values[i] for i in range(min(len(keys), len(values)))}
        if any(_clean(v) for k, v in cells.items() if k != "district_name"):
            used.append((ROW_FIRST_DATA + offset, values))

    if len(used) > MAX_ROWS:
        sheet_errors.append(f"The file has {len(used)} farmers — the limit is {MAX_ROWS} per "
                            f"upload. Please split it.")
        return [], [], sheet_errors

    for row_no, values in used:
        cells = {keys[i]: values[i] for i in range(min(len(keys), len(values)))}
        err = RowErrors(row_no)
        payload = _row_to_payload(cells, vocab, by_label, by_id, stap_by_activity,
                                  district_id, err)

        mobile = _clean(cells.get("mobile_no"))
        if mobile:
            if mobile in known_mobiles:
                err.add(f"Mobile {mobile} is already registered")
            if mobile in seen_mobiles:
                err.add(f"Mobile {mobile} is also on row {seen_mobiles[mobile]}")
            else:
                seen_mobiles[mobile] = row_no

        digits = payload.get("aadhaar_no")
        if digits:
            if aadhaar_hash(digits) in known_aadhaar:
                err.add("This Aadhaar is already registered")
            if digits in seen_aadhaar:
                err.add(f"The same Aadhaar is also on row {seen_aadhaar[digits]}")
            else:
                seen_aadhaar[digits] = row_no

        if err:
            errors.append({"row": row_no, "name": _row_label(cells),
                           "errors": err.messages, "raw": list(values)})
        else:
            ready.append(payload)

    return ready, errors, sheet_errors


def _row_label(cells: dict) -> str:
    return " ".join(x for x in (_clean(cells.get("first_name")),
                                _clean(cells.get("last_name"))) if x) or "(no name)"


def _output_staps_by_activity(db: Session, activity_ids: list[str]) -> dict[str, list[str]]:
    """Ticking an activity selects every OUTPUT row beneath it — the same expansion
    StapGroupPicker performs, so a bulk-created farmer matches a form-created one."""
    out: dict[str, list[str]] = {}
    rows = db.query(SilkTypeActivityProduct).filter(
        SilkTypeActivityProduct.activity_id.in_(activity_ids or [""]),
        SilkTypeActivityProduct.role == "OUTPUT").all()
    for row in rows:
        out.setdefault(row.activity_id, []).append(row.id)
    return out


def _row_to_payload(cells: dict, vocab: dict, by_label: dict, by_id: dict,
                    stap_by_activity: dict, district_id: str, err: RowErrors) -> dict:
    payload: dict[str, Any] = {"district_id": district_id}

    if _clean(cells.get("district_name")) != vocab["district_name"]:
        err.add(f"District must be {vocab['district_name']} — you are acting as that district. "
                f"Download a fresh template if you meant a different one.")

    for col in COLUMNS:
        if col.key in ("district_name", "date_of_birth", "aadhaar_no", LAC_KEY,
                       "experience_activity_ids", "land_type", "dag_no", "patta_no",
                       "asset_type_id", "asset_quantity", "asset_year"):
            continue
        raw = _clean(cells.get(col.key))
        if not raw:
            if col.required:
                err.add(f"{col.header} is required")
            continue
        if col.source == "Genders" and raw not in GENDERS:
            err.add(f"Gender must be one of {', '.join(GENDERS)}")
        elif col.source == "FarmerTypes" and raw not in FARMER_TYPES:
            err.add(f"Farmer Type must be one of {', '.join(FARMER_TYPES)}")
        elif col.source == "Circles":
            resolved = vocab["circles"].get(raw)
            if not resolved:
                err.add(f"Sericulture Circle '{raw}' is not in your district")
            payload[col.key] = resolved
            continue
        elif col.source in ("Castes", "Religions", "EducationLevels"):
            table = {"Castes": "castes", "Religions": "religions",
                     "EducationLevels": "education"}[col.source]
            resolved = vocab[table].get(raw)
            if not resolved:
                err.add(f"{col.header} '{raw}' is not a known value")
            payload[col.key] = resolved
            continue
        if col.key in ("experience_years", "family_member_male", "family_member_female"):
            number = _parse_int(raw)
            if number is None or number < 0:
                err.add(f"{col.header} must be a whole number")
            else:
                payload[col.key] = number
            continue
        payload[col.key] = raw

    _dob(cells, payload, err)
    _aadhaar(cells, payload, err)
    _experience_activities(cells, by_label, payload, err)
    _production_activities(cells, by_id, stap_by_activity, payload, err)
    _land(cells, payload, err)
    _asset(cells, vocab, payload, err)
    return payload


def _dob(cells: dict, payload: dict, err: RowErrors) -> None:
    raw = cells.get("date_of_birth")
    if not _clean(raw):
        err.add("Date of Birth is required")
        return
    parsed = _parse_date(raw)
    if not parsed:
        err.add("Date of Birth must be dd-mm-yyyy")
    elif parsed >= date.today():
        err.add("Date of Birth must be in the past")
    else:
        payload["date_of_birth"] = parsed


def _aadhaar(cells: dict, payload: dict, err: RowErrors) -> None:
    raw = _clean(cells.get("aadhaar_no"))
    if not raw:
        err.add("Aadhaar is required")
        return
    try:
        payload["aadhaar_no"] = normalize_aadhaar(raw)
    except ValueError as exc:
        err.add(str(exc))


def _experience_activities(cells: dict, by_label: dict, payload: dict, err: RowErrors) -> None:
    raw = _clean(cells.get("experience_activity_ids"))
    if not raw:
        err.add("Experience in Activities is required")
        payload["experience_activity_ids"] = []
        return
    ids: list[str] = []
    for part in (p.strip() for p in raw.split(";")):
        if not part:
            continue
        activity = by_label.get(part)
        if not activity:
            err.add(f"'{part}' is not in the Activities reference list")
        elif activity.id not in ids:
            ids.append(activity.id)
    payload["experience_activity_ids"] = ids


def _production_activities(cells: dict, by_id: dict, stap_by_activity: dict,
                           payload: dict, err: RowErrors) -> None:
    stap_ids: list[str] = []
    picked = 0
    for key, value in cells.items():
        if not key.startswith("activity:"):
            continue
        answer = _clean(value).lower()
        if answer in ("", "no"):
            continue
        if answer != "yes":
            err.add(f"'{_clean(value)}' is not valid for an activity column — use Yes or No")
            continue
        picked += 1
        activity_id = key.split(":", 1)[1]
        for stap_id in stap_by_activity.get(activity_id, []):
            if stap_id not in stap_ids:
                stap_ids.append(stap_id)
        if activity_id in by_id and not stap_by_activity.get(activity_id):
            err.add(f"'{by_id[activity_id].activity_name}' has no output product configured yet")
    if not picked:
        err.add("Tick Yes against at least one silk type / activity")
    payload["stap_ids"] = stap_ids


def _land(cells: dict, payload: dict, err: RowErrors) -> None:
    dag, patta = _clean(cells.get("dag_no")), _clean(cells.get("patta_no"))
    land_type = _clean(cells.get("land_type"))
    if not (dag or patta or land_type):
        payload["lands"] = []
        return
    if land_type and land_type not in VALID_LAND_TYPES:
        err.add(f"Land Type must be one of {', '.join(sorted(VALID_LAND_TYPES))}")
        land_type = ""
    payload["lands"] = [{"dag_no": dag, "patta_no": patta, "land_type": land_type or "Owned"}]


def _asset(cells: dict, vocab: dict, payload: dict, err: RowErrors) -> None:
    name = _clean(cells.get("asset_type_id"))
    quantity_raw = _clean(cells.get("asset_quantity"))
    year_raw = _clean(cells.get("asset_year"))
    if not name:
        if quantity_raw or year_raw:
            err.add("Quantity / Year Acquired need an Asset Type")
        payload["assets"] = []
        return
    asset_type_id = vocab["asset_types"].get(name)
    if not asset_type_id:
        err.add(f"Asset Type '{name}' is not available for an individual farmer")
        payload["assets"] = []
        return
    quantity = _parse_int(quantity_raw) if quantity_raw else 1
    if quantity is None or quantity < 1:
        err.add("Quantity must be a whole number of 1 or more")
        quantity = 1
    year = None
    if year_raw:
        year = _parse_int(year_raw)
        if year is None or not (1900 <= year <= date.today().year):
            err.add(f"Year Acquired must be a year between 1900 and {date.today().year}")
            year = None
    payload["assets"] = [{"asset_type_id": asset_type_id, "quantity": quantity,
                          "acquisition_year": year}]


# --------------------------------------------------------------------------- error report

def errors_to_xlsx(errors: list[dict], sheet_errors: list[str]) -> bytes:
    """The uploaded rows with a trailing Errors column, so the admin fixes in place."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Errors"
    headers = ["Row", "Farmer", "What to fix"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for item in errors:
        ws.append([item["row"], item["name"], "; ".join(item["errors"])])
    for item in sheet_errors:
        ws.append(["—", "(whole file)", item])
    for width, letter in zip((8, 28, 120), ("A", "B", "C")):
        ws.column_dimensions[letter].width = width
    for row in ws.iter_rows(min_row=2):
        row[2].alignment = Alignment(wrap_text=True, vertical="top")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
